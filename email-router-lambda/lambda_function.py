import json
import boto3
import io
from typing import List, Dict

from openpyxl import load_workbook
from strands import Agent, tool

# =========================
# CONFIG: CHANGE THESE
# =========================

S3_BUCKET = "lambda-email-router-role"
S3_KEY = "projects.xlsx"
DEFAULT_SUPERVISOR_EMAIL = "charithdhanam@gmail.com"
SES_REGION = "us-east-1"
FROM_EMAIL = "forwardinginbox99@gmail.com"

# =========================
# AWS CLIENTS
# =========================

s3_client = boto3.client("s3")
ses_client = boto3.client("ses", region_name=SES_REGION)

# =========================
# TOOLS FOR THE AGENT
# =========================

@tool
def load_projects_from_excel() -> List[Dict]:
    """
    Load project info from an Excel file stored in S3.

    Expected columns in each sheet:
    ProjectName | SupervisorName | SupervisorEmail | Keywords
    """
    obj = s3_client.get_object(Bucket=S3_BUCKET, Key=S3_KEY)
    data = obj["Body"].read()

    wb = load_workbook(io.BytesIO(data), data_only=True)

    projects: List[Dict] = []

    for sheet in wb.worksheets:
        # first row = header
        header = {}
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if isinstance(cell_value, str):
                header[cell_value.strip()] = col

        required = ["ProjectName", "SupervisorEmail", "Keywords"]
        if not all(name in header for name in required):
            continue

        row = 2
        while row <= sheet.max_row:
            project_name = sheet.cell(row=row, column=header["ProjectName"]).value
            supervisor_email = sheet.cell(row=row, column=header["SupervisorEmail"]).value
            keywords = sheet.cell(row=row, column=header["Keywords"]).value

            if project_name and supervisor_email:
                projects.append(
                    {
                        "project_name": str(project_name).strip(),
                        "supervisor_email": str(supervisor_email).strip(),
                        "keywords": str(keywords or "").strip(),
                    }
                )
            row += 1

    return projects


@tool
def send_supervisor_email(
    supervisor_email: str,
    original_from: str,
    original_subject: str,
    project_name: str,
    email_summary: str,
) -> str:
    """
    Send a notification email to the supervisor via SES.
    """
    subject = f"[Project: {project_name}] New email received"
    body_text = (
        f"Dear Supervisor,\n\n"
        f"You have a new email related to your project: {project_name}.\n\n"
        f"From: {original_from}\n"
        f"Original subject: {original_subject}\n\n"
        f"Summary of the email:\n{email_summary}\n\n"
        f"Please check the original email in the shared mailbox.\n\n"
        f"Regards,\nProject Router"
    )

    response = ses_client.send_email(
        Source=FROM_EMAIL,
        Destination={"ToAddresses": [supervisor_email]},
        Message={
            "Subject": {"Data": subject},
            "Body": {"Text": {"Data": body_text}},
        },
    )

    return f"Email sent to {supervisor_email} with MessageId {response['MessageId']}"


# =========================
# STRANDS AGENT
# =========================

system_prompt = """
You are a routing assistant for project emails.

Steps:
1. Call load_projects_from_excel to get all projects.
2. Read the email subject and body.
3. Decide which ONE project best matches, using its name and keywords.
4. If no project matches, use the fallback supervisor email.
5. Create a short 2-4 sentence summary of the email.
6. Call send_supervisor_email once with:
   - supervisor_email (chosen or fallback)
   - original_from
   - original_subject
   - project_name (or "Unassigned")
   - email_summary

Return JSON with:
- project_name
- supervisor_email
- summary
"""

agent = Agent(
    tools=[load_projects_from_excel, send_supervisor_email],
)



# =========================
# SES EVENT PARSING (simple)
# =========================

def extract_email_from_ses_event(event):
    """
    Get From and Subject from SES commonHeaders.
    For now, body is a placeholder.
    """
    record = event["Records"][0]
    mail = record["ses"]["mail"]
    common = mail.get("commonHeaders", {})

    from_list = common.get("from", ["unknown@example.com"])
    from_address = from_list[0]
    subject = common.get("subject", "(no subject)")

    body_text = "(Email body parsing not implemented yet.)"

    return from_address, subject, body_text


# =========================
# LAMBDA HANDLER
# =========================

def lambda_handler(event, context):
    print("Received event:", json.dumps(event))

    try:
        original_from, subject, body_text = extract_email_from_ses_event(event)

        user_message = (
            f"New email received.\n\n"
            f"From: {original_from}\n"
            f"Subject: {subject}\n\n"
            f"Body:\n{body_text}\n"
        )

        # Combine our system instructions + this specific email
        full_prompt = system_prompt + "\n\n" + user_message

        prompt = (
            f"Original email:\n"
            f"From: {original_from}\n"
            f"Subject: {subject}\n"
            f"Body: {body_text}\n"
        )

        result = agent(prompt)

        print("Agent result:", result)

        return {
            "statusCode": 200,
            "body": json.dumps(
                {
                    "message": "Processed email",
                    "agent_result": str(result),
                }
            ),
        }

    except Exception as e:
        print("Error:", str(e))
        return {
            "statusCode": 500,
            "body": json.dumps({"error": str(e)}),
        }
