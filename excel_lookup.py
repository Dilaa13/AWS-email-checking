from __future__ import annotations

import io
import os

import boto3
from openpyxl import load_workbook

S3_BUCKET = os.environ.get("PROJECTS_BUCKET", "my-project-config-bucket")
S3_KEY = os.environ.get("PROJECTS_KEY", "config/projects.xlsx")

s3_client = boto3.client("s3")


def _load_workbook_from_s3():
    """Download the workbook from S3 into memory and return an openpyxl workbook."""
    resp = s3_client.get_object(Bucket=S3_BUCKET, Key=S3_KEY)
    data = resp["Body"].read()
    return load_workbook(filename=io.BytesIO(data), data_only=True)


def get_supervisor_email(project_sheet_name: str) -> str | None:
    """
    Given a project sheet name (e.g. 'e-commerce', 'AI-project', 'SE-project'),
    return the SupervisorEmail from that sheet, or None if not found.
    """

    wb = _load_workbook_from_s3()

    if project_sheet_name not in wb.sheetnames:
        return None

    ws = wb[project_sheet_name]

    # Find which column is 'SupervisorEmail' by reading the header row (row 1)
    header_cells = ws[1]  # first row
    header_map = {cell.value: cell.column for cell in header_cells}

    email_col = header_map.get("SupervisorEmail")
    if email_col is None:
        return None

    # For now we assume the supervisor info is in row 2
    supervisor_email = ws.cell(row=2, column=email_col).value
    return supervisor_email
